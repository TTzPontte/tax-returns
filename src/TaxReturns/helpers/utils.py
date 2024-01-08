# %%
# def parse_to_brl(value):
#     return "R$ {:,.2f}".format(value)
import re

from .aztronic import get_client


def get_client_email(cnpj_cpf):
    email = get_client(cnpj_cpf)
    return email['cliente']['email']


# def parse_to_brl(num):
#     if num == 0:
#         return "R$ 0,00"
#     parts = str(num).split('.')
#     int_part = parts[0][::-1]
#     dec_part = parts[1][::-1] if len(parts) > 1 else None
#     formatted_int_part = ".".join([int_part[i:i + 3][::-1] for i in range(0, len(int_part), 3)])[::-1]
#     formatted_number = formatted_int_part + ("," + dec_part if dec_part else "")
#     return 'R$ ' + str(formatted_number)
def parse_to_brl(f):
    if f == 0:
        return "R$ 0,00"
    else:
        return 'R$ ' + '{:,.2f}'.format(f)


def format_br_doc(doc_number):
    doc_number = ''.join(c for c in doc_number if c.isdigit())
    if len(doc_number) == 11:
        formatted_doc = '{}.{}.{}-{}'.format(doc_number[:3], doc_number[3:6], doc_number[6:9], doc_number[9:11])
        return formatted_doc
    elif len(doc_number) == 14:
        formatted_doc = '{}.{}.{}/{}-{}'.format(doc_number[:2], doc_number[2:5], doc_number[5:8], doc_number[8:12],
                                                doc_number[12:14])
        return formatted_doc
    else:
        return "Invalid document number."


def is_valid_email(email):
    email_regex = re.compile(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')
    return bool(email_regex.match(email))


# 1000000.00  -> 1, 000, 000.00
# 1, 234, 609.45  -> 1234609.45


# %%
import json
from datetime import datetime
from numbers import Number

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# sys.path.append(os.path.join("..", "PorttalUsuario", "remover_acesso", "Relação_contratos.xlsx"))
def build_date(date):
    if isinstance(date, Number):
        return datetime.fromtimestamp(date / 1000).strftime('%Y-%m-%dT%H:%M:%SZ')
    else:
        return datetime.fromisoformat(date).strftime('%Y-%m-%dT%H:%M:%SZ')


def csv_to_json(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                column_letter_plus_row = ws[column_letter + str(1)].value
                column_title = column_letter + str(row)
                column_value = ws[column_title].value
                if isinstance(column_value, datetime):
                    print("column_value", column_value)
                    column_value = build_date(str(column_value))
                # print("column_letter_plus_row",column_letter_plus_row)
                # print("column_title",column_title)
                # print("column_value",column_value)
                my_dict[column_letter_plus_row] = column_value
        my_list.append(my_dict)

    data = json.dumps(my_list, sort_keys=True, indent=4)  # with open('D:/data.json', 'w', encoding='utf-8') as f:
    return json.loads(data)


def do_it():
    data = csv_to_json('/Users/Mr-i-me/code/Mr-i-me-pontte/portal-scripts/layers/helpers/CARTORIOS DA BASE PONTTE.xlsx')

    with open('/Users/Mr-i-me/code/Mr-i-me-pontte/portal-scripts/layers/helpers/data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
